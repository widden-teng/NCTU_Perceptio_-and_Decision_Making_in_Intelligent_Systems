import numpy as np
from PIL import Image
import numpy as np
import habitat_sim
from habitat_sim.utils.common import d3_40_colors_rgb
import cv2
import json
import math
import openpyxl
import os
from scipy.io import loadmat

# This is the scene we are going to load.
# support a variety of mesh formats, such as .glb, .gltf, .obj, .ply
### put your scene path ###
test_scene = "/home/widden/NCTU/Perception_and_Decision_Making/dataset/apartment_0/habitat/mesh_semantic.ply"
path = "/home/widden/NCTU/Perception_and_Decision_Making/dataset/apartment_0/habitat/info_semantic.json"

# global test_pic
# instance id to semantic id
with open(path, "r") as f:
    annotations = json.load(f)
# load colors
colors = loadmat('color101.mat')['colors']
colors = np.insert(colors, 0, values=np.array(
    [[0, 0, 0]]), axis=0)  # to make the color be correct

id_to_label = []
instance_id_to_semantic_label_id = np.array(annotations["id_to_label"])
for i in instance_id_to_semantic_label_id:
    if i < 0:
        id_to_label.append(0)
    else:
        id_to_label.append(i)
id_to_label = np.asarray(id_to_label)

######

sim_settings = {
    "scene": test_scene,  # Scene path
    "default_agent": 0,  # Index of the default agent
    "sensor_height": 1.5,  # Height of sensors in meters, relative to the agent
    "width": 512,  # Spatial resolution of the observations
    "height": 512,
    "sensor_pitch": 0,  # sensor pitch (x rotation in rads)
}

# This function generates a config for the simulator.
# It contains two parts:
# one for the simulator backend
# one for the agent, where you can attach a bunch of sensors


def transform_rgb_bgr(image):
    return image[:, :, [2, 1, 0]]


def transform_depth(image):
    depth_img = (image / 10 * 255).astype(np.uint8)
    return depth_img


def transform_semantic(semantic_obs):
    semantic_img = Image.new(
        "P", (semantic_obs.shape[1], semantic_obs.shape[0]))
    semantic_img.putpalette(colors.flatten())
    semantic_img.putdata(semantic_obs.flatten().astype(np.uint8))
    semantic_img = semantic_img.convert("RGB")
    semantic_img = cv2.cvtColor(np.asarray(semantic_img), cv2.COLOR_RGB2BGR)
    return semantic_img


def make_simple_cfg(settings, dis_per_step, agl_per_step):
    # simulator backend
    sim_cfg = habitat_sim.SimulatorConfiguration()
    sim_cfg.scene_id = settings["scene"]

    # In the 1st example, we attach only one sensor,
    # a RGB visual sensor, to the agent
    rgb_sensor_spec = habitat_sim.CameraSensorSpec()
    rgb_sensor_spec.uuid = "color_sensor"
    rgb_sensor_spec.sensor_type = habitat_sim.SensorType.COLOR
    rgb_sensor_spec.resolution = [settings["height"], settings["width"]]
    rgb_sensor_spec.position = [0.0, settings["sensor_height"], 0.0]
    rgb_sensor_spec.orientation = [
        settings["sensor_pitch"],
        0.0,
        0.0,
    ]
    rgb_sensor_spec.sensor_subtype = habitat_sim.SensorSubType.PINHOLE

    # depth snesor
    depth_sensor_spec = habitat_sim.CameraSensorSpec()
    depth_sensor_spec.uuid = "depth_sensor"
    depth_sensor_spec.sensor_type = habitat_sim.SensorType.DEPTH
    depth_sensor_spec.resolution = [settings["height"], settings["width"]]
    depth_sensor_spec.position = [0.0, settings["sensor_height"], 0.0]
    depth_sensor_spec.orientation = [
        settings["sensor_pitch"],
        0.0,
        0.0,
    ]
    depth_sensor_spec.sensor_subtype = habitat_sim.SensorSubType.PINHOLE

    # semantic snesor
    semantic_sensor_spec = habitat_sim.CameraSensorSpec()
    semantic_sensor_spec.uuid = "semantic_sensor"
    semantic_sensor_spec.sensor_type = habitat_sim.SensorType.SEMANTIC
    semantic_sensor_spec.resolution = [settings["height"], settings["width"]]
    semantic_sensor_spec.position = [0.0, settings["sensor_height"], 0.0]
    semantic_sensor_spec.orientation = [
        settings["sensor_pitch"],
        0.0,
        0.0,
    ]
    semantic_sensor_spec.sensor_subtype = habitat_sim.SensorSubType.PINHOLE

    # agent
    agent_cfg = habitat_sim.agent.AgentConfiguration()

    agent_cfg.sensor_specifications = [
        rgb_sensor_spec, depth_sensor_spec, semantic_sensor_spec]
    ##################################################################
    # change the move_forward length or rotate angle
    ##################################################################
    agent_cfg.action_space = {
        "move_forward": habitat_sim.agent.ActionSpec(
            "move_forward", habitat_sim.agent.ActuationSpec(
                amount=dis_per_step)  # 0.01 means 0.01 m
        ),
        "turn_left": habitat_sim.agent.ActionSpec(
            "turn_left", habitat_sim.agent.ActuationSpec(
                amount=agl_per_step)  # 1.0 means 1 degree
        ),
        "turn_right": habitat_sim.agent.ActionSpec(
            "turn_right", habitat_sim.agent.ActuationSpec(amount=1.0)
        ),
    }

    return habitat_sim.Configuration(sim_cfg, [agent_cfg])


# initialize parament
path_nodes = np.load("path_nodes.npy")
dis_per_step = 0.01
agl_per_step = 1.0

cfg = make_simple_cfg(sim_settings, dis_per_step, agl_per_step)
sim = habitat_sim.Simulator(cfg)


# initialize an agent
agent = sim.initialize_agent(sim_settings["default_agent"])

# Set agent state
agent_state = habitat_sim.AgentState()
agent_state.position = np.array(
    [path_nodes[0][0], 0.0, path_nodes[0][1] + 0.01])  # agent in world space
agent.set_state(agent_state)

# obtain the default, discrete actions that an agent can perform
# default action space contains 3 actions: move_forward, turn_left, and turn_right
action_names = list(
    cfg.agents[sim_settings["default_agent"]].action_space.keys())
print("Discrete action space: ", action_names)


FORWARD_KEY = "w"
LEFT_KEY = "a"
RIGHT_KEY = "d"
FINISH = "f"
STEP_PLUS_ONE = "q"
print("#############################")
print("use keyboard to control the agent")
print(" w for go forward  ")
print(" a for turn left  ")
print(" d for trun right  ")
print(" f for finish and quit the program")
print("#############################")


def object_highlight(rgb, sem):
    index = np.where((sem[:, :, 2] == object_color[0]) & (
        sem[:, :, 1] == object_color[1]) & (sem[:, :, 0] == object_color[2]))
    if len(index[0]) != 0:
        rgb[index] = cv2.addWeighted(rgb[index], 0.6, sem[index], 0.4, 50)
    return rgb


def navigateAndSee(action=""):
    global posit
    if action in action_names:
        observations = sim.step(action)
        # print("action: ", action)
        highlighted = object_highlight(transform_rgb_bgr(
            observations["color_sensor"]), transform_semantic(id_to_label[observations["semantic_sensor"]]))
        cv2.imshow("RGB", highlighted)
        # cv2.imshow("depth", transform_depth(observations["depth_sensor"]))
        cv2.imshow("semantic", transform_semantic(
            id_to_label[observations["semantic_sensor"]]))
        agent_state = agent.get_state()
        sensor_state = agent_state.sensor_states['color_sensor']
        # print("camera pose: x y z rw rx ry rz")
        # print(sensor_state.position[0], sensor_state.position[1], sensor_state.position[2],
        #   sensor_state.rotation.w, sensor_state.rotation.x, sensor_state.rotation.y, sensor_state.rotation.z)
        posit = [sensor_state.position[0],
                 sensor_state.position[1], sensor_state.position[2]]
        return highlighted


def point_dif(current_face, last_face):
    theta = np.rad2deg(np.arccos(current_face.dot(
        last_face)/(np.linalg.norm(current_face) * np.linalg.norm(last_face))))
    if np.isnan(theta):
        theta = 90
    return theta, np.linalg.norm(current_face)


# def rotate_direction():
#     cross = np.cross(last_face, current_face)

object_dict = {}
object_name = ['refrigerator', 'rack', 'cushion', 'lamp', 'cooktop']
wb = openpyxl.load_workbook(
    'color_coding_semantic_segmentation_classes.xlsx')
s1 = wb['Sheet1']
for i in range(1, s1.max_row + 1):
    if s1.cell(i, 5).value in object_name:
        temp = s1.cell(i, 2).value.replace(
            "(", "").replace(")", "").replace(",", "")
        temp = np.array([int(tp)for tp in temp.split()])
        object_dict[s1.cell(i, 5).value] = temp
# target = "lamp"
target = input("input target :")
object_color = object_dict[target]

if not (os.path.exists("./video")):
    os.makedirs('video')
# save video initial
path = "video/" + target + ".mp4"
fourcc = cv2.VideoWriter_fourcc(*'mp4v')
videowriter = cv2.VideoWriter(path, fourcc, 100, (512, 512))

action = "move_forward"
action_img = navigateAndSee(action)
videowriter.write(action_img)

arrive = False
Angle_temp_arrive = False
Distance_temp_arrive = False
node_num = 0
last_face = np.array([0, -1])
current_face = []
last_node = path_nodes[node_num]
current_node = [posit[0], posit[2]]
new_node = path_nodes[node_num+1]
theta = 0
dis_step = 0
agl_step = 0
temp_step = 0
while not arrive:
    keystroke = cv2.waitKey(0)
    if keystroke == ord(STEP_PLUS_ONE):
        current_face = new_node - current_node
        theta, distance = point_dif(current_face, last_face)
        # 先轉角度
        if agl_step != int(theta / agl_per_step) and (not Angle_temp_arrive):
            print("total ratate step :", int(theta / agl_per_step))
            if np.cross(last_face, current_face) > 0:
                action = "turn_right"
            else:
                action = "turn_left"
            agl_step += 1
            print("angle_approuching")
            print("i th turn : ", agl_step)
        elif not Angle_temp_arrive:
            Angle_temp_arrive = not Angle_temp_arrive
            agl_step = 0
            print("angle_finish")
        # 在直走
        if Angle_temp_arrive and (dis_step != int(distance / dis_per_step)):
            action = "move_forward"
            dis_step += 1
            print("total walk step :", int(distance / dis_per_step))
            print("i th move foward : ", dis_step)
        elif Angle_temp_arrive and (not Distance_temp_arrive):
            print("straight_finish")
            Distance_temp_arrive = not Distance_temp_arrive

        if Distance_temp_arrive & Angle_temp_arrive:
            print("{}th node is find".format(node_num))
            dis_step = 0
            node_num += 1
            print(node_num)
            print(len(path_nodes))
            Angle_temp_arrive = not Angle_temp_arrive
            Distance_temp_arrive = not Distance_temp_arrive
            last_node = current_node.copy()
            current_node = [posit[0], posit[2]]
            new_node = path_nodes[node_num+1]
            # last_node = path_nodes[node_num-1]
            last_face = [current_node[i] - last_node[i] for i in range(2)]
            print("ideal point : ", path_nodes[node_num])
            print("current  point : ", current_node)
            continue
        else:
            action_img = navigateAndSee(action)
            videowriter.write(action_img)

    elif keystroke == ord(FINISH):
        print("action: FINISH")
        break
    else:
        print("INVALID KEY")
        continue
    if node_num == len(path_nodes)-2:  # 這邊很奇怪, 本來是2
        arrive = not arrive
        videowriter.release()
        print("finish  ")
        break
print("find the target point!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!")
