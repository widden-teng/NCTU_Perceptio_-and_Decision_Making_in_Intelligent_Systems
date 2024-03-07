import cv2
import numpy as np
import math
import random
import argparse
import os
import openpyxl


class Nodes:
    """Class to store the RRT graph"""

    def __init__(self, x, y):
        self.x = x
        self.y = y
        self.parent_x = []
        self.parent_y = []

# check collision


def collision(x1, y1, x2, y2):
    color = []
    # 在兩點間直線取100間隔去判斷是否有碰到障礙物
    if (x1 != x2):
        x = list(np.arange(x1, x2, (x2-x1)/100))
        y = list(((y2-y1)/(x2-x1))*(x-x1) + y1)
        # print("collision", x, y)
        for i in range(len(x)):
            # 因為code中的x, y方向與img相反, 所以會相反
            color.append(img[int(y[i]), int(x[i])])
        if (0 in color):  # 若有一個全黑
            return True  # collision
        else:
            return False  # no-collision
    else:
        x = x1
        y = list(np.arange(y1, y2, (y2-y1)/100))
        # print("collision", x, y)
        for i in range(len(y)):
            # 因為code中的x, y方向與img相反, 所以會相反
            color.append(img[int(y[i]), int(x[i])])
        if (0 in color):  # 若有一個全黑
            return True  # collision
        else:
            return False  # no-collision


# check the  collision with obstacle and trim
# stepSize 這邊才會用到
def check_collision(x1, y1, x2, y2, end_points):
    _, theta = dist_and_angle(x2, y2, x1, y1)
    x = x2 + stepSize*np.cos(theta)
    y = y2 + stepSize*np.sin(theta)
    # print(x2, y2, x1, y1)
    # print("theta", theta)
    # print("check_collision", x, y)

    # TODO: trim the branch if its going out of image area
    # print("Image shape",img.shape)
    hy, hx = img.shape
    directCon = False
    nodeCon = False
    if y < 0 or y > hy or x < 0 or x > hx:
        # print("Point out of image bound")
        {}
    else:
        # check connection between two nodes
        if collision(x, y, x2, y2):
            nodeCon = False
        else:
            nodeCon = True
            # 是否該node(與給定點多一個step size的點）與終點相連
            if collision(x, y,  end_points[0], end_points[1]):
                directCon = False
            else:
                directCon = True

    return (x, y, directCon, nodeCon)


# return dist and angle b/w new point and nearest node
def dist_and_angle(x1, y1, x2, y2):
    dist = math.sqrt(((x1-x2)**2)+((y1-y2)**2))
    angle = math.atan2(y2-y1, x2-x1)
    return (dist, angle)


# return the neaerst node index
def nearest_node(x, y, temp_node_list):
    temp_dist = []
    for i in range(len(temp_node_list)):
        # _為忽略回傳值
        dist, _ = dist_and_angle(
            x, y, temp_node_list[i].x, temp_node_list[i].y)
        temp_dist.append(dist)
    return temp_dist.index(min(temp_dist))

# generate a random point in the image space

# 回傳 0~369  0~496 隨機值


def rnd_point(h, l):
    new_y = random.randint(0, h)
    new_x = random.randint(0, l)
    return (new_x, new_y)


def RRT(img, img2, start, end, stepSize):
    h, l = img.shape  # dim of the loaded image

    # insert the starting point in the node class
    # node_list = [0] # 裡面存有所有node的位置資訊
    # 起點長的
    temp_node_list = [0]
    end_node_list = [0]
    temp_node_list[0] = Nodes(start[0], start[1])
    temp_node_list[0].parent_x.append(start[0])
    temp_node_list[0].parent_y.append(start[1])

    # 終點長的
    end_node_list[0] = Nodes(end[0], end[1])
    end_node_list[0].parent_x.append(end[0])
    end_node_list[0].parent_y.append(end[1])

    # display start and end , 都為藍色圈圈
    cv2.circle(img2, (start[0], start[1]), 5,
               (0, 0, 255), thickness=3, lineType=8)
    cv2.circle(img2, (end[0], end[1]), 5, (0, 0, 255), thickness=3, lineType=8)
    end_points = [end[0], end[1]]
    i = 1
    num_nodes = 1
    pathFound = False
    while pathFound == False:
        nx, ny = rnd_point(h, l)
        nearest_ind = nearest_node(nx, ny, temp_node_list)
        nearest_x = temp_node_list[nearest_ind].x
        nearest_y = temp_node_list[nearest_ind].y

        outerside_nearestid = nearest_node(nx, ny, end_node_list)
        end_points = [end_node_list[outerside_nearestid].x,
                      end_node_list[outerside_nearestid].y]
        # check direct connection
        # 會往前走一個step tx, ty
        tx, ty, directCon, nodeCon = check_collision(
            nx, ny, nearest_x, nearest_y, end_points)
        # print("Check collision:", tx, ty, directCon, nodeCon)

        # 都true表示可以直接與終點相連
        if directCon and nodeCon:
            # print("Node can connect directly with end")
            k = len(temp_node_list)
            temp_node_list.append(k)  # 用於初始化node_list[i] ,沒其他功能
            temp_node_list[k] = Nodes(tx, ty)
            # list.copy() 用於list內容的複製，而非記憶體複製
            # 此處將最近點的路徑都複製進新走一步的parent list 中

            temp_node_list[k].parent_x = temp_node_list[nearest_ind].parent_x.copy()
            temp_node_list[k].parent_y = temp_node_list[nearest_ind].parent_y.copy()
            temp_node_list[k].parent_x.append(tx)
            temp_node_list[k].parent_y.append(ty)

            for j in range(len(end_node_list[outerside_nearestid].parent_x)):
                temp_node_list[k].parent_x.append(
                    end_node_list[outerside_nearestid].parent_x[-(j+1)])
                temp_node_list[k].parent_y.append(
                    end_node_list[outerside_nearestid].parent_y[-(j+1)])

            cv2.circle(img2, (int(tx), int(ty)), 2,
                       (0, 0, 255), thickness=3, lineType=8)
            # 將最近點與剛走得step以綠色相連
            cv2.line(img2, (int(tx), int(ty)), (int(temp_node_list[nearest_ind].x), int(
                temp_node_list[nearest_ind].y)), (0, 255, 0), thickness=1, lineType=8)

            # 最佳相連的路徑變藍色
            for j in range(len(temp_node_list[k].parent_x)-1):
                cv2.line(img2, (int(temp_node_list[k].parent_x[j]), int(temp_node_list[k].parent_y[j])), (int(
                    temp_node_list[k].parent_x[j+1]), int(temp_node_list[k].parent_y[j+1])), (255, 0, 0), thickness=2, lineType=8)
            cv2.waitKey(1)
            cv2.imwrite("media/"+str(i)+".jpg", img2)
            cv2.imwrite("out.jpg", img2)
            cv2.imshow("out.jpg", img2)
            cv2.waitKey()
            break

        # 表示新的step不會撞牆, 但還沒到終點
        elif nodeCon:

            temp_node_list.append(i)
            j = len(temp_node_list)-1
            temp_node_list[j] = Nodes(tx, ty)
            temp_node_list[j].parent_x = temp_node_list[nearest_ind].parent_x.copy()
            temp_node_list[j].parent_y = temp_node_list[nearest_ind].parent_y.copy()
            temp_node_list[j].parent_x.append(tx)
            temp_node_list[j].parent_y.append(ty)

            print("{}th iteration".format(i))
            i = i+1
            num_nodes = num_nodes + 1

            # display
            cv2.circle(img2, (int(tx), int(ty)), 2,
                       (0, 0, 255), thickness=3, lineType=8)
            # 將最近點與剛走得step以綠色相連
            cv2.line(img2, (int(tx), int(ty)), (int(temp_node_list[nearest_ind].x), int(
                temp_node_list[nearest_ind].y)), (0, 255, 0), thickness=1, lineType=8)
            cv2.imwrite("media/"+str(i)+".jpg", img2)
            cv2.imshow("sdc", img2)
            cv2.waitKey(1)
            temp_list = temp_node_list.copy()
            temp_node_list = end_node_list.copy()
            end_node_list = temp_list.copy()
            continue

        else:
            # print("No direct con. and no node con. :( Generating new rnd numbers")
            continue

    if (temp_node_list[k].parent_x[0] != start[0]):
        temp_node_list[k].parent_x.reverse()
        temp_node_list[k].parent_y.reverse()
        print("reverse")
    temp_x = temp_node_list[k].parent_x
    temp_y = temp_node_list[k].parent_y
    temp_x = np.float64(temp_x)*(17/l) - 6
    temp_y = 7 - (np.float64(temp_y)*(11/h))
    final_nodes = np.empty((0, 2), float)
    for j in range(temp_x.shape[0]):
        final_nodes = np.append(final_nodes, [[temp_y[j], temp_x[j]]], axis=0)
    np.save("path_nodes", final_nodes)


def draw_circle(event, x, y, flags, param):
    global coordinates
    if event == cv2.EVENT_LBUTTONDBLCLK:
        cv2.circle(img2, (x, y), 5, (255, 0, 0), -1)
        coordinates.append(x)
        coordinates.append(y)


def find_target_point(img, img2):
    object_dict = {}
    object_name = ['refrigerator', 'rack', 'cushion', 'lamp', 'cooktop']
    wb = openpyxl.load_workbook(
        'color_coding_semantic_segmentation_classes.xlsx')
    s1 = wb['Sheet1']
    for i in range(1, s1.max_row + 1):
        if s1.cell(i, 5).value in object_name:
            temp = s1.cell(i, 2).value.replace(
                "(", "").replace(")", "").replace(",", "")
            temp = np.array([int(tp)for tp in temp.split()])
            object_dict[s1.cell(i, 5).value] = temp
    img2 = np.array(img2)
    flag = 0

    for k in object_name:
        index = np.where((img2[:, :, 2] == object_dict[k][0]) & (
            img2[:, :, 1] == object_dict[k][1]) & (img2[:, :, 0] == object_dict[k][2]))
        temp_x = int(np.round(np.mean(index[0])))
        temp_y = int(np.round(np.mean(index[1])))
        flag = 1
        while True:
            up = [temp_x, temp_y+flag]
            down = [temp_x, temp_y-flag]
            left = [temp_x-flag, temp_y]
            right = [temp_x+flag, temp_y]
            if img[up[0], up[1]] == 255:
                xy_points = up
                break
            elif img[down[0], down[1]] == 255:
                xy_points = down
                break
            elif img[left[0], left[1]] == 255:
                xy_points = left
                break
            elif img[right[0], right[1]] == 255:
                xy_points = right
                break
            else:
                flag = flag + 1
        object_dict[k] = [xy_points[1], xy_points[0]]
        print(object_dict[k])
        if (k == 'refrigerator'):
            object_dict[k] = [175, 195]
        elif (k == 'cushion'):
            object_dict[k] = [400, 200]
        elif (k == 'cooktop'):
            object_dict[k] = [144, 233]

    return (object_dict)
    print(object_dict)


if __name__ == '__main__':

    parser = argparse.ArgumentParser(description='Below are the params:')
    parser.add_argument('-p', type=str, default='map.png', metavar='ImagePath', action='store', dest='imagePath',
                        help='Path of the image containing mazes')
    parser.add_argument('-s', type=int, default=20, metavar='Stepsize', action='store', dest='stepSize',
                        help='Step-size to be used for RRT branches')
    parser.add_argument('-start', type=int, default=[360, 199], metavar='startCoord', dest='start', nargs='+',
                        help='Starting position in the maze')
    parser.add_argument('-stop', type=int, default=[372, 302], metavar='stopCoord', dest='stop', nargs='+',
                        help='End position in the maze')
    parser.add_argument(
        '-selectPoint', help='Select start and end points from figure', action='store_true')
    parser.add_argument(
        '-demo', help='give target and select start point', action='store_true')

    args = parser.parse_args()

    # 刪除media資料夾
    try:
        os.system("rm -rf media")
    except:
        {
            # print("Dir already clean")
        }
    os.mkdir("media")

    img = cv2.imread(args.imagePath, 0)  # load grayscale maze image
    # 將灰階圖轉黑白圖
    thresh = 230
    img = cv2.threshold(img, thresh, 255, cv2.THRESH_BINARY)[1]
    kernel = np.ones((5, 5), np.uint8)
    img = cv2.erode(img, kernel, iterations=2)
    cv2.imshow("erosion", img)
    cv2.waitKey()
    img2 = cv2.imread(args.imagePath)  # 全彩圖
    start = tuple(args.start)  # (20,20) # starting coordinate
    end = tuple(args.stop)  # (450,250) # target coordinate
    stepSize = args.stepSize  # stepsize for RRT
    temp_node_list = [0]  # list to store all the node points 順便初始化
    end_node_list = [0]
    coordinates = []
    object_dict = find_target_point(img, img2)
    # 判斷點哪裡
    if args.selectPoint:

        cv2.namedWindow('image')
        cv2.setMouseCallback('image', draw_circle)
        while (1):
            cv2.imshow('image', img2)
            k = cv2.waitKey(20) & 0xFF
            if k == 27:
                break

        start = (coordinates[0], coordinates[1])
        end = (coordinates[2], coordinates[3])
    if args.demo:

        cv2.namedWindow('image')
        cv2.setMouseCallback('image', draw_circle)
        while (1):
            cv2.imshow('image', img2)
            k = cv2.waitKey(20) & 0xFF
            if k == 27 or len(coordinates) == 2:
                break

        start = (coordinates[0], coordinates[1])
        target = input("input target :")
        if target in object_dict:
            end = (object_dict[target][0], object_dict[target][1])
    # run the RRT algorithm
    RRT(img, img2, start, end, stepSize)
